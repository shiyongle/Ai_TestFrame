"""
知识库文件导入工具
支持多格式文件解析并生成文档内容
"""

from typing import Dict, Any, List, Tuple
import io
import os
import logging

logger = logging.getLogger(__name__)


def _as_text(data: bytes, encoding: str = "utf-8") -> str:
    try:
        return data.decode(encoding)
    except UnicodeDecodeError:
        return data.decode("utf-8", errors="ignore")


def _extract_markdown(file_name: str, data: bytes) -> Tuple[str, str]:
    return os.path.splitext(file_name)[0], _as_text(data)


def _extract_docx(file_name: str, data: bytes) -> Tuple[str, str]:
    try:
        from docx import Document  # type: ignore
    except Exception as exc:
        raise RuntimeError("缺少依赖: python-docx") from exc

    doc = Document(io.BytesIO(data))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return os.path.splitext(file_name)[0], "\n".join(paragraphs)


def _extract_xlsx(file_name: str, data: bytes) -> Tuple[str, str]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("缺少依赖: openpyxl") from exc

    wb = load_workbook(filename=io.BytesIO(data), data_only=True)
    parts: List[str] = []
    for sheet in wb.worksheets:
        parts.append(f"[Sheet] {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell) for cell in row if cell is not None]
            if row_values:
                parts.append(" | ".join(row_values))
    return os.path.splitext(file_name)[0], "\n".join(parts)


def _extract_xmind(file_name: str, data: bytes) -> Tuple[str, str]:
    try:
        from xmindparser import xmind_to_dict  # type: ignore
    except Exception as exc:
        raise RuntimeError("缺少依赖: xmindparser") from exc

    temp_path = f"{os.path.splitext(file_name)[0]}.xmind"
    with open(temp_path, "wb") as f:
        f.write(data)
    try:
        mind = xmind_to_dict(temp_path)
    finally:
        try:
            os.remove(temp_path)
        except OSError:
            pass

    def walk(node: Dict[str, Any], depth: int = 0, lines: List[str] = None) -> List[str]:
        lines = lines or []
        title = node.get("title")
        if title:
            lines.append("  " * depth + f"- {title}")
        for child in node.get("topics", []) or []:
            walk(child, depth + 1, lines)
        return lines

    lines: List[str] = []
    for sheet in mind or []:
        if "topic" in sheet:
            walk(sheet["topic"], 0, lines)
    return os.path.splitext(file_name)[0], "\n".join(lines)


def parse_knowledge_file(file_name: str, data: bytes) -> Dict[str, str]:
    """解析文件为知识文档内容"""
    ext = os.path.splitext(file_name)[1].lower()
    if ext in [".md", ".markdown", ".txt"]:
        title, content = _extract_markdown(file_name, data)
    elif ext in [".docx"]:
        title, content = _extract_docx(file_name, data)
    elif ext in [".xlsx", ".xls"]:
        title, content = _extract_xlsx(file_name, data)
    elif ext in [".xmind"]:
        title, content = _extract_xmind(file_name, data)
    else:
        raise RuntimeError(f"暂不支持该格式: {ext}")

    return {"title": title, "content": content}


def parse_multiple_files(files: List[Tuple[str, bytes]]) -> List[Dict[str, str]]:
    """批量解析文件"""
    results = []
    logger.info(f"开始批量解析 {len(files)} 个文件")
    
    for file_name, data in files:
        try:
            logger.info(f"解析文件: {file_name}, 大小: {len(data)} 字节")
            result = parse_knowledge_file(file_name, data)
            logger.info(f"成功解析文件: {file_name}")
            results.append(result)
        except Exception as exc:
            logger.error(f"解析文件失败: {file_name} - {exc}")
            import traceback
            traceback.print_exc()
            # 不抛出异常，继续处理其他文件
            results.append({"title": file_name, "content": f"解析失败: {str(exc)}", "error": True})
    
    logger.info(f"批量解析完成，成功: {len([r for r in results if not r.get('error')])}, 失败: {len([r for r in results if r.get('error')])}")
    return results
