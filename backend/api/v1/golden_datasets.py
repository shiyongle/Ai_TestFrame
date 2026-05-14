"""
黄金测试集 API 路由
"""

from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session

from api.deps import get_database
from schemas.agent_evaluation_schemas import (
    GoldenDatasetCreate,
    GoldenDatasetItemCreate,
    GoldenDatasetItemUpdate,
    GoldenDatasetUpdate,
)
from services.golden_dataset_service import golden_dataset_service

router = APIRouter()


@router.get("/golden-datasets")
async def list_golden_datasets(
    keyword: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_database),
):
    datasets, total = golden_dataset_service.list_datasets(db, keyword=keyword, limit=limit, offset=offset)
    return {
        "total": total,
        "items": [golden_dataset_service.serialize_dataset(d) for d in datasets],
    }


@router.post("/golden-datasets")
async def create_golden_dataset(
    payload: GoldenDatasetCreate,
    db: Session = Depends(get_database),
):
    dataset = golden_dataset_service.create_dataset(db, payload)
    return golden_dataset_service.serialize_dataset(dataset, include_items=True)


# ============ Excel 模板下载（必须在 {dataset_id} 路由之前注册） ============

@router.get("/golden-datasets/template/download")
async def download_excel_template():
    """下载黄金测试集 Excel 导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "黄金测试集"

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    sample_font = Font(name="微软雅黑", size=11, color="666666", italic=True)
    note_font = Font(name="微软雅黑", size=10, color="999999")

    # 表头
    headers = [
        ("问题（必填）", 45),
        ("期望答案（必填）", 55),
        ("分类（可选）", 18),
        ("优先级（可选）", 15),
        ("标签（可选，多个用逗号分隔）", 30),
    ]
    for col_idx, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # 示例数据
    samples = [
        ("你们的退货政策是什么？", "我们支持7天无理由退货，商品需保持完好包装。", "退货政策", "high", "客服,售后"),
        ("如何查询物流信息？", '您可以在订单详情页点击"物流跟踪"查看实时物流状态。', "物流查询", "medium", "客服"),
        ("会员积分怎么兑换？", '登录后进入"我的积分"页面，选择兑换商品即可。', "会员权益", "low", "会员"),
    ]
    for row_idx, row_data in enumerate(samples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = sample_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 说明行
    ws.cell(row=6, column=1, value="【填写说明】").font = Font(name="微软雅黑", bold=True, size=10, color="E74C3C")
    notes = [
        "1. 「问题」和「期望答案」为必填列，其他列可选",
        "2. 「优先级」可填 high / medium / low，默认 medium",
        "3. 「标签」多个标签用英文逗号分隔，如：客服,售后",
        "4. 请从第 2 行开始填写数据，第 1 行为表头请勿修改",
        "5. 上方灰色斜体行是示例，导入前请删除或覆盖",
    ]
    for i, note in enumerate(notes):
        ws.cell(row=7 + i, column=1, value=note).font = note_font

    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=golden_dataset_template.xlsx"},
    )


# ============ {dataset_id} 路由 ============

@router.get("/golden-datasets/{dataset_id}")
async def get_golden_dataset(
    dataset_id: int,
    db: Session = Depends(get_database),
):
    dataset = golden_dataset_service.get_dataset(db, dataset_id)
    if not dataset:
        raise HTTPException(status_code=404, detail="黄金测试集不存在")
    return golden_dataset_service.serialize_dataset(dataset, include_items=True)


@router.put("/golden-datasets/{dataset_id}")
async def update_golden_dataset(
    dataset_id: int,
    payload: GoldenDatasetUpdate,
    db: Session = Depends(get_database),
):
    dataset = golden_dataset_service.update_dataset(db, dataset_id, payload)
    if not dataset:
        raise HTTPException(status_code=404, detail="黄金测试集不存在")
    return golden_dataset_service.serialize_dataset(dataset)


@router.delete("/golden-datasets/{dataset_id}")
async def delete_golden_dataset(
    dataset_id: int,
    db: Session = Depends(get_database),
):
    ok = golden_dataset_service.delete_dataset(db, dataset_id)
    if not ok:
        raise HTTPException(status_code=404, detail="黄金测试集不存在")
    return {"message": "已删除"}


@router.post("/golden-datasets/{dataset_id}/items")
async def add_golden_dataset_items(
    dataset_id: int,
    items: List[GoldenDatasetItemCreate],
    db: Session = Depends(get_database),
):
    try:
        created = golden_dataset_service.add_items(db, dataset_id, items)
        return [golden_dataset_service.serialize_item(i) for i in created]
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/golden-datasets/{dataset_id}/import-excel")
async def import_excel(
    dataset_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_database),
):
    """从 Excel 文件导入测试条目到指定黄金测试集"""
    from openpyxl import load_workbook
    import io

    # 校验文件类型
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

    # 校验测试集存在
    dataset = golden_dataset_service.get_dataset(db, dataset_id)
    if not dataset:
        raise HTTPException(status_code=404, detail="黄金测试集不存在")

    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 文件解析失败: {str(e)}")

    items_to_create = []
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 2:
            continue
        question = str(row[0] or "").strip()
        expected_answer = str(row[1] or "").strip()

        # 跳过空行和说明行
        if not question or not expected_answer:
            if question and not expected_answer:
                errors.append(f"第{row_idx}行：缺少期望答案")
            continue
        if question.startswith("【") or question.startswith("1.") or question.startswith("2."):
            continue  # 跳过说明文字

        category = str(row[2] or "").strip() if len(row) > 2 else ""
        priority = str(row[3] or "medium").strip().lower() if len(row) > 3 else "medium"
        tags_str = str(row[4] or "").strip() if len(row) > 4 else ""

        if priority not in ("high", "medium", "low"):
            priority = "medium"

        tags = [t.strip() for t in tags_str.split(",") if t.strip()] if tags_str else []

        items_to_create.append(GoldenDatasetItemCreate(
            question=question,
            expected_answer=expected_answer,
            category=category or None,
            priority=priority,
            tags=tags or None,
        ))

    if not items_to_create:
        raise HTTPException(
            status_code=400,
            detail=f"Excel 中未找到有效数据。{'发现错误: ' + '; '.join(errors) if errors else '请确认从第2行开始填写，且问题和期望答案均不为空'}",
        )

    try:
        created = golden_dataset_service.add_items(db, dataset_id, items_to_create)
        return {
            "message": f"成功导入 {len(created)} 条数据",
            "imported_count": len(created),
            "error_count": len(errors),
            "errors": errors[:10] if errors else [],
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ============ 单条条目操作 ============

@router.put("/golden-dataset-items/{item_id}")
async def update_golden_dataset_item(
    item_id: int,
    payload: GoldenDatasetItemUpdate,
    db: Session = Depends(get_database),
):
    item = golden_dataset_service.update_item(db, item_id, payload)
    if not item:
        raise HTTPException(status_code=404, detail="条目不存在")
    return golden_dataset_service.serialize_item(item)


@router.delete("/golden-dataset-items/{item_id}")
async def delete_golden_dataset_item(
    item_id: int,
    db: Session = Depends(get_database),
):
    ok = golden_dataset_service.delete_item(db, item_id)
    if not ok:
        raise HTTPException(status_code=404, detail="条目不存在")
    return {"message": "已删除"}
